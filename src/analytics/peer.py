"""Peer Comparison Engine — Day 18 deliverable.

DAD-PROJ-001 Sprint 3.

Computes PERCENT_RANK for 10 metrics within each of 11 peer groups,
populates the peer_percentiles table in SQLite, and generates
peer_comparison.xlsx with 11 colour-coded sheets.
"""

import logging
import os
import sqlite3
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, os.path.join(PROJECT_ROOT, "src", "screener"))

logger = logging.getLogger(__name__)

DB_PATH = os.path.join(PROJECT_ROOT, "db", "nifty100.db")
SUPPORTING_DIR = os.path.join(PROJECT_ROOT, "data", "supporting")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")

# 10 metrics to rank within each peer group
PEER_METRICS = [
    "return_on_equity_pct",
    "net_profit_margin_pct",
    "debt_to_equity",          # inverse — lower is better
    "free_cash_flow_cr",
    "asset_turnover",
    "interest_coverage",
    "earnings_per_share",
    "dividend_yield_pct",
    "pe_ratio",                # inverse — lower is better (value)
    "market_cap_crore",
]

INVERSE_METRICS = {"debt_to_equity", "pe_ratio"}  # lower = better rank

# Colour fills for Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def load_peer_data() -> tuple:
    """Load peer groups and financial ratios data."""
    pg = pd.read_excel(os.path.join(SUPPORTING_DIR, "peer_groups.xlsx"), header=0)
    pg["company_id"] = pg["company_id"].astype(str).str.strip().str.upper()

    fr = pd.read_excel(os.path.join(SUPPORTING_DIR, "financial_ratios.xlsx"), header=0)
    fr["company_id"] = fr["company_id"].astype(str).str.strip().str.upper()
    fr = fr.sort_values("year").groupby("company_id").last().reset_index()

    mc = pd.read_excel(os.path.join(SUPPORTING_DIR, "market_cap.xlsx"), header=0)
    mc["company_id"] = mc["company_id"].astype(str).str.strip().str.upper()
    mc = mc.sort_values("year").groupby("company_id").last().reset_index()
    mc = mc[["company_id", "market_cap_crore", "pe_ratio", "pb_ratio", "dividend_yield_pct"]]

    companies = pd.read_excel(
        os.path.join(PROJECT_ROOT, "data", "raw", "companies.xlsx"), header=1
    )
    companies["id"] = companies["id"].astype(str).str.strip().str.upper()
    companies = companies[["id", "company_name"]].rename(columns={"id": "company_id"})

    data = fr.merge(mc, on="company_id", how="left", suffixes=("", "_mc"))
    data = data.merge(companies, on="company_id", how="left")

    return pg, data


def compute_percentile_rank(series: pd.Series, inverse: bool = False) -> pd.Series:
    """Compute PERCENT_RANK within a group (0 to 1 scale).

    Args:
        series: Values to rank.
        inverse: If True, lower value = higher rank (e.g. D/E, P/E).

    Returns:
        Percentile ranks as a Series (0 = lowest, 1 = highest).
    """
    if inverse:
        return 1 - series.rank(pct=True, na_option="bottom")
    return series.rank(pct=True, na_option="bottom")


def compute_peer_percentiles(pg: pd.DataFrame, data: pd.DataFrame) -> pd.DataFrame:
    """Compute percentile ranks for all companies within their peer groups.

    Returns:
        Long-format DataFrame with columns:
        company_id, peer_group_name, metric, value, percentile_rank, year
    """
    records = []

    for group_name in pg["peer_group_name"].unique():
        group_members = pg[pg["peer_group_name"] == group_name]["company_id"].tolist()
        group_data = data[data["company_id"].isin(group_members)].copy()

        if group_data.empty:
            continue

        for metric in PEER_METRICS:
            if metric not in group_data.columns:
                continue

            is_inv = metric in INVERSE_METRICS
            ranks = compute_percentile_rank(group_data[metric], inverse=is_inv)

            for idx, row in group_data.iterrows():
                records.append({
                    "company_id": row["company_id"],
                    "peer_group_name": group_name,
                    "metric": metric,
                    "value": row.get(metric),
                    "percentile_rank": round(ranks[idx], 4),
                    "year": row.get("year"),
                })

    return pd.DataFrame(records)


def write_peer_percentiles_to_db(percentiles_df: pd.DataFrame) -> int:
    """Write peer percentiles to SQLite peer_percentiles table."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    # Create table if it doesn't exist
    conn.execute("""
        CREATE TABLE IF NOT EXISTS peer_percentiles (
            id INTEGER PRIMARY KEY,
            company_id TEXT NOT NULL,
            peer_group_name TEXT NOT NULL,
            metric TEXT NOT NULL,
            value REAL,
            percentile_rank REAL,
            year TEXT,
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)
    conn.execute("DELETE FROM peer_percentiles")
    conn.commit()

    # Only write companies that exist in companies table
    valid_ids = set(
        row[0] for row in conn.execute("SELECT id FROM companies").fetchall()
    )
    filtered = percentiles_df[percentiles_df["company_id"].isin(valid_ids)]
    filtered.to_sql("peer_percentiles", conn, if_exists="append", index=False)
    conn.commit()

    count = conn.execute("SELECT COUNT(*) FROM peer_percentiles").fetchone()[0]
    conn.close()
    return count


def _percentile_fill(pct: float) -> PatternFill:
    """Green >= 75th, Yellow 25th-75th, Red <= 25th percentile."""
    if pct is None:
        return None
    if pct >= 0.75:
        return GREEN_FILL
    elif pct >= 0.25:
        return YELLOW_FILL
    return RED_FILL


def generate_peer_comparison_excel(
    pg: pd.DataFrame,
    data: pd.DataFrame,
    percentiles_df: pd.DataFrame,
    output_path: str,
) -> None:
    """Generate peer_comparison.xlsx with 11 sheets — one per peer group."""
    wb = Workbook()
    wb.remove(wb.active)

    for group_name in sorted(pg["peer_group_name"].unique()):
        group_members = pg[pg["peer_group_name"] == group_name]
        benchmark_ids = set(
            group_members[group_members["is_benchmark"] == 1]["company_id"]
        )
        member_ids = group_members["company_id"].tolist()

        group_data = data[data["company_id"].isin(member_ids)].copy()
        if group_data.empty:
            continue

        ws = wb.create_sheet(title=group_name[:31])

        # Build display columns
        display_cols = ["company_id", "company_name"] + [
            m for m in PEER_METRICS if m in group_data.columns
        ]
        group_data = group_data[display_cols].copy()

        # Add percentile rank columns
        group_pct = percentiles_df[
            percentiles_df["peer_group_name"] == group_name
        ].pivot(index="company_id", columns="metric", values="percentile_rank")
        group_pct.columns = [f"{c}_pct" for c in group_pct.columns]
        group_data = group_data.merge(
            group_pct.reset_index(), on="company_id", how="left"
        )

        # Header row
        ws.append([group_name] + [""] * (len(group_data.columns) - 1))
        ws.append(list(group_data.columns))

        # Style header
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for cell in ws[2]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        for _, row in group_data.iterrows():
            row_vals = []
            for col in group_data.columns:
                val = row[col]
                if isinstance(val, float) and not pd.isna(val):
                    val = round(val, 2)
                elif not isinstance(val, str) and pd.isna(val):
                    val = None
                row_vals.append(val)
            ws.append(row_vals)

            # Colour-code percentile cells + benchmark row
            data_row = ws.max_row
            is_benchmark = row["company_id"] in benchmark_ids

            for col_idx, col_name in enumerate(group_data.columns, start=1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

                if is_benchmark:
                    cell.fill = GOLD_FILL
                elif col_name.endswith("_pct"):
                    fill = _percentile_fill(cell.value)
                    if fill:
                        cell.fill = fill

        # Summary row — peer group median
        ws.append([])
        median_row = ["MEDIAN", ""]
        for col in display_cols[2:]:
            if col in group_data.columns:
                vals = group_data[col].dropna()
                median_row.append(round(vals.median(), 2) if not vals.empty else None)
        ws.append(median_row)

        ws.freeze_panes = "C3"

        # Auto-width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 20)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info("Saved peer_comparison.xlsx to %s", output_path)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s"
    )

    print("Loading peer data...")
    pg, data = load_peer_data()

    print("Computing percentile rankings...")
    percentiles_df = compute_peer_percentiles(pg, data)
    print(f"Computed {len(percentiles_df)} percentile records across {percentiles_df['peer_group_name'].nunique()} groups")

    print("Writing to SQLite...")
    count = write_peer_percentiles_to_db(percentiles_df)
    print(f"peer_percentiles table: {count} rows")

    output_path = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")
    print("Generating peer_comparison.xlsx...")
    generate_peer_comparison_excel(pg, data, percentiles_df, output_path)
    print(f"Done — saved to {output_path}")

    print("\nPeer group summary:")
    for grp in sorted(pg["peer_group_name"].unique()):
        members = pg[pg["peer_group_name"] == grp]["company_id"].tolist()
        print(f"  {grp}: {len(members)} companies")
