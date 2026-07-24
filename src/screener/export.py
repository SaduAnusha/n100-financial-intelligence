"""Screener Export — Day 17 deliverable.

Generates output/screener_output.xlsx with one sheet per preset screener.
Each sheet contains 20 KPI columns sorted by composite_quality_score,
with colour-coded cells (green = meets threshold, red = fails threshold).

Run directly: python src/screener/export.py
"""

import logging
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, HERE)
sys.path.insert(0, os.path.join(PROJECT_ROOT, "src", "analytics"))
sys.path.insert(0, os.path.join(PROJECT_ROOT, "src", "etl"))

from engine import load_screener_data, load_config, compute_composite_score, run_preset  # noqa

logger = logging.getLogger(__name__)

OUTPUT_PATH = os.path.join(PROJECT_ROOT, "output", "screener_output.xlsx")

# Columns to include in the Excel output (20 KPI columns)
OUTPUT_COLUMNS = [
    "company_id", "company_name", "broad_sector",
    "composite_quality_score",
    "return_on_equity_pct", "debt_to_equity", "free_cash_flow_cr",
    "net_profit_margin_pct", "operating_profit_margin_pct",
    "asset_turnover", "interest_coverage",
    "earnings_per_share", "book_value_per_share",
    "dividend_payout_ratio_pct", "dividend_yield_pct",
    "pe_ratio", "pb_ratio", "market_cap_crore",
    "sales", "net_profit",
    "cash_from_operations_cr", "total_debt_cr",
]

# Colour fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Threshold checks per column — (operator, threshold)
THRESHOLD_CHECKS = {
    "return_on_equity_pct": ("min", 15.0),
    "debt_to_equity": ("max", 1.0),
    "free_cash_flow_cr": ("min", 0.0),
    "net_profit_margin_pct": ("min", 10.0),
    "operating_profit_margin_pct": ("min", 15.0),
    "interest_coverage": ("min", 3.0),
    "dividend_yield_pct": ("min", 2.0),
    "pe_ratio": ("max", 30.0),
    "pb_ratio": ("max", 4.0),
}


def _cell_color(col: str, value) -> PatternFill:
    """Return green if value meets threshold, red if it fails, None otherwise."""
    if col not in THRESHOLD_CHECKS or value is None:
        return None
    operator, threshold = THRESHOLD_CHECKS[col]
    try:
        val = float(value)
    except (TypeError, ValueError):
        return None
    if operator == "min":
        return GREEN_FILL if val >= threshold else RED_FILL
    elif operator == "max":
        return GREEN_FILL if val <= threshold else RED_FILL
    return None


def write_preset_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, preset_config: dict) -> None:
    """Write one preset's results to a worksheet."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars

    # Title row
    ws.append([f"Preset: {sheet_name} — {preset_config.get('description', '')}"])
    ws.append([f"Companies: {len(df)}  |  Expected range: {preset_config.get('expected_count', 'N/A')}"])
    ws.append([])  # blank row

    # Get available columns
    avail_cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
    ws.append(avail_cols)

    # Style title rows
    for row_idx in [1, 2]:
        for cell in ws[row_idx]:
            cell.fill = TITLE_FILL
            cell.font = TITLE_FONT

    # Style header row (row 4)
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for _, row in df[avail_cols].iterrows():
        row_data = []
        for col in avail_cols:
            val = row[col]
            if isinstance(val, float) and not pd.isna(val):
                val = round(val, 2)
            elif pd.isna(val) if not isinstance(val, str) else False:
                val = None
            row_data.append(val)
        ws.append(row_data)

        # Colour-code data cells
        data_row = ws.max_row
        for col_idx, col_name in enumerate(avail_cols, start=1):
            cell = ws.cell(row=data_row, column=col_idx)
            fill = _cell_color(col_name, cell.value)
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)

    ws.freeze_panes = "A5"


def generate_screener_excel(output_path: str = OUTPUT_PATH) -> None:
    """Generate screener_output.xlsx with 6 preset sheets."""
    config = load_config()
    df = load_screener_data()
    df["composite_quality_score"] = compute_composite_score(df, config)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    for preset_name, preset_config in config["presets"].items():
        result = run_preset(df, preset_name, config)
        write_preset_sheet(wb, preset_name, result, preset_config)
        logger.info("Sheet '%s': %d rows written.", preset_name, len(result))

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info("Saved screener_output.xlsx to %s", output_path)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s"
    )
    print("Generating screener_output.xlsx...")
    generate_screener_excel()
    print(f"Done — saved to {OUTPUT_PATH}")
